from io import BytesIO
from pathlib import Path
import sys

from openpyxl import load_workbook

sys.path.append(str(Path(__file__).resolve().parents[1]))

from app.routes.parse import ParseRequest, export_campaign_excel


def test_export_excel_endpoint_returns_xlsx_file():
    response = export_campaign_excel(
        ParseRequest(raw_text="https://example.com", campaign_name="Lumma Stealer")
    )

    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert response.headers["content-disposition"] == "attachment; filename=defender_iocs.xlsx"
    assert response.body.startswith(b"PK")


def test_export_excel_endpoint_excludes_sender_email_rows():
    response = export_campaign_excel(
        ParseRequest(raw_text="https://example.com user@test.com")
    )

    workbook = load_workbook(filename=BytesIO(response.body))
    worksheet = workbook.active

    values = [worksheet[row][1].value for row in range(2, worksheet.max_row + 1)]
    assert "https://example.com" in values
    assert "user@test.com" not in values
