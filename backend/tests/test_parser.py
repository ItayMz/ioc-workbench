from pathlib import Path
import sys
from io import BytesIO

from openpyxl import Workbook

sys.path.append(str(Path(__file__).resolve().parents[1]))

from app.enums.action import Action
from app.enums.category import Category
from app.enums.indicator_type import IndicatorType
from app.services.parser import parse_bulk_text, parse_ioc, prepare_text_from_upload


def test_parse_ioc_returns_enum_indicator_for_url():
    parsed = parse_ioc("https://example.com")

    assert parsed.indicator_type is IndicatorType.URL
    assert parsed.valid is True


def test_parse_ioc_returns_enum_indicator_for_ip_and_domain():
    ip_ioc = parse_ioc("8.8.8.8")
    domain_ioc = parse_ioc("example.com")

    assert ip_ioc.indicator_type is IndicatorType.IP_ADDRESS
    assert domain_ioc.indicator_type is IndicatorType.DOMAIN_NAME


def test_parse_ioc_returns_enum_indicator_for_email_and_hashes():
    email_ioc = parse_ioc("user@example.com")
    md5_ioc = parse_ioc("0123456789abcdef0123456789abcdef")
    sha1_ioc = parse_ioc("aaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa")
    sha256_ioc = parse_ioc("aaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa")

    assert email_ioc.indicator_type is IndicatorType.SENDER_EMAIL_ADDRESS
    assert md5_ioc.indicator_type is IndicatorType.FILE_MD5
    assert sha1_ioc.indicator_type is IndicatorType.FILE_SHA1
    assert sha256_ioc.indicator_type is IndicatorType.FILE_SHA256


def test_hash_indicators_get_block_and_remediate_action():
    md5_ioc = parse_ioc("0123456789abcdef0123456789abcdef")
    sha1_ioc = parse_ioc("aaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa")
    sha256_ioc = parse_ioc("aaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa")

    assert md5_ioc.action is Action.BLOCK_AND_REMEDIATE
    assert sha1_ioc.action is Action.BLOCK_AND_REMEDIATE
    assert sha256_ioc.action is Action.BLOCK_AND_REMEDIATE


def test_non_hash_indicators_get_block_action_and_invalid_have_none():
    url_ioc = parse_ioc("https://example.com")
    ip_ioc = parse_ioc("8.8.8.8")
    domain_ioc = parse_ioc("example.com")
    email_ioc = parse_ioc("user@example.com")
    invalid_ioc = parse_ioc("not-an-ioc")

    assert url_ioc.action is Action.BLOCK
    assert ip_ioc.action is Action.BLOCK
    assert domain_ioc.action is Action.BLOCK
    assert email_ioc.action is Action.BLOCK
    assert invalid_ioc.action is None


def test_valid_indicators_get_malware_category_and_alerts():
    url_ioc = parse_ioc("https://example.com")
    hash_ioc = parse_ioc("0123456789abcdef0123456789abcdef")

    assert url_ioc.category is Category.MALWARE
    assert url_ioc.generate_alert is True
    assert hash_ioc.category is Category.MALWARE
    assert hash_ioc.generate_alert is True


def test_invalid_indicators_have_no_category_or_alerts():
    invalid_ioc = parse_ioc("not-an-ioc")

    assert invalid_ioc.category is None
    assert invalid_ioc.generate_alert is None


def test_valid_indicators_get_severity_and_expiration_defaults():
    url_ioc = parse_ioc("https://example.com")
    hash_ioc = parse_ioc("0123456789abcdef0123456789abcdef")

    assert url_ioc.severity == "High"
    assert url_ioc.expiration_time == "2099-12-31T23:59:59.0Z"
    assert hash_ioc.severity == "High"
    assert hash_ioc.expiration_time == "2099-12-31T23:59:59.0Z"


def test_invalid_indicators_have_no_severity_or_expiration_defaults():
    invalid_ioc = parse_ioc("not-an-ioc")

    assert invalid_ioc.severity is None
    assert invalid_ioc.expiration_time is None


def test_parse_bulk_text_handles_empty_and_whitespace_input():
    assert parse_bulk_text("") == []
    assert parse_bulk_text("   \n\t  ") == []


def test_parse_bulk_text_trims_whitespace_and_ignores_blank_lines():
    indicators = parse_bulk_text("\n  https://example.com   \n\n  8.8.8.8  \n")

    assert [indicator.refanged_value for indicator in indicators] == ["https://example.com", "8.8.8.8"]
    assert [indicator.indicator_type for indicator in indicators] == [IndicatorType.URL, IndicatorType.IP_ADDRESS]


def test_parse_bulk_text_ignores_email_addresses_completely():
    indicators = parse_bulk_text("user@test.com\n1.1.1.1\nevil.com")

    assert [indicator.refanged_value for indicator in indicators] == ["1.1.1.1", "evil.com"]
    assert [indicator.indicator_type for indicator in indicators] == [IndicatorType.IP_ADDRESS, IndicatorType.DOMAIN_NAME]


def test_parse_ioc_rejects_extremely_long_input_without_crashing():
    parsed = parse_ioc("x" * 20000)

    assert parsed.valid is False
    assert parsed.reason == "line_too_long"


def test_prepare_text_from_upload_rejects_empty_files():
    try:
        prepare_text_from_upload(b"", "test.csv")
    except ValueError as exc:
        assert "empty" in str(exc).lower()
    else:
        raise AssertionError("Expected ValueError for empty upload")


def test_prepare_text_from_upload_rejects_unsupported_file_types():
    try:
        prepare_text_from_upload(b"https://example.com", "test.exe")
    except ValueError as exc:
        assert "unsupported" in str(exc).lower()
    else:
        raise AssertionError("Expected ValueError for unsupported file type")


def test_prepare_text_from_upload_rejects_invalid_utf8():
    try:
        prepare_text_from_upload(b"\xff\xfe\x00", "test.csv")
    except ValueError as exc:
        assert "utf-8" in str(exc).lower()
    else:
        raise AssertionError("Expected ValueError for invalid UTF-8")


def test_prepare_text_from_upload_rejects_malformed_csv():
    try:
        prepare_text_from_upload(b"https://example.com,8.8.8.8\nhttps://example.com", "test.csv")
    except ValueError as exc:
        assert "malformed csv" in str(exc).lower()
    else:
        raise AssertionError("Expected ValueError for malformed CSV")


def test_prepare_text_from_upload_accepts_xlsx_with_blank_rows_and_mixed_types():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["8.8.8.8"])
    worksheet.append([None])
    worksheet.append([12345])
    worksheet.append(["evil.com"])

    buffer = BytesIO()
    workbook.save(buffer)

    prepared = prepare_text_from_upload(buffer.getvalue(), "sample.xlsx")
    assert prepared == "8.8.8.8\n12345\nevil.com"


def test_prepare_text_from_upload_rejects_corrupted_xlsx():
    try:
        prepare_text_from_upload(b"not-a-valid-workbook", "bad.xlsx")
    except ValueError as exc:
        assert "xlsx" in str(exc).lower() and "invalid" in str(exc).lower()
    else:
        raise AssertionError("Expected ValueError for corrupted XLSX")


def test_prepare_text_from_upload_reads_multiple_xlsx_sheets_and_ignores_blank_rows():
    workbook = Workbook()
    first = workbook.active
    first.title = "SheetA"
    first.append(["https://one.example"])
    first.append([None])

    second = workbook.create_sheet(title="SheetB")
    second.append(["8.8.8.8"])

    third = workbook.create_sheet(title="Blank")
    third.append([None])

    buffer = BytesIO()
    workbook.save(buffer)

    prepared = prepare_text_from_upload(buffer.getvalue(), "multi.xlsx")
    assert prepared == "https://one.example\n8.8.8.8"
