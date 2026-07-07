from io import BytesIO
from pathlib import Path
import sys

from openpyxl import load_workbook

sys.path.append(str(Path(__file__).resolve().parents[1]))

from app.enums.action import Action
from app.enums.category import Category
from app.enums.indicator_type import IndicatorType
from app.exporters.defender_excel import HEADERS, export_campaign_to_excel_bytes
from app.models.campaign import Campaign
from app.models.ioc import ParsedIOC


def test_exporter_writes_expected_headers_and_filters_invalid_indicators():
    valid_indicator = ParsedIOC(
        original_value="https://example.com",
        refanged_value="https://example.com",
        indicator_type=IndicatorType.URL,
        action=Action.BLOCK,
        category=Category.MALWARE,
        generate_alert=True,
        severity="High",
        expiration_time="2099-12-31T23:59:59.0Z",
        valid=True,
    )
    invalid_indicator = ParsedIOC(
        original_value="not-an-ioc",
        refanged_value="not-an-ioc",
        valid=False,
    )
    campaign = Campaign(
        campaign_name="Lumma Stealer",
        title="Block Lumma Stealer Indicators",
        description="Indicators associated with the Lumma Stealer campaign.",
        recommended_actions="Block the listed indicators and investigate any historical communication.",
        indicators=[valid_indicator, invalid_indicator],
    )

    workbook_bytes = export_campaign_to_excel_bytes(campaign)
    workbook = load_workbook(filename=BytesIO(workbook_bytes))
    worksheet = workbook.active

    assert worksheet.max_row == 2
    assert worksheet.max_column == len(HEADERS)
    assert [cell.value for cell in worksheet[1]] == HEADERS
    assert worksheet[2][0].value == "Url"
    assert worksheet[2][1].value == "https://example.com"
    assert worksheet[2][2].value == "2099-12-31T23:59:59.0Z"
    assert worksheet[2][3].value == "Block"
    assert worksheet[2][4].value == "High"
    assert worksheet[2][5].value == "Block Lumma Stealer Indicators"
    assert worksheet[2][6].value == "Indicators associated with the Lumma Stealer campaign."
    assert worksheet[2][7].value == "Block the listed indicators and investigate any historical communication."
    assert worksheet[2][8].value is None
    assert worksheet[2][9].value == "Malware"
    assert worksheet[2][10].value is None
    assert worksheet[2][11].value is True


def test_exporter_writes_hash_action_as_block_and_remediate():
    indicator = ParsedIOC(
        original_value="0123456789abcdef0123456789abcdef",
        refanged_value="0123456789abcdef0123456789abcdef",
        indicator_type=IndicatorType.FILE_MD5,
        action=Action.BLOCK_AND_REMEDIATE,
        category=Category.MALWARE,
        generate_alert=True,
        severity="High",
        expiration_time="2099-12-31T23:59:59.0Z",
        valid=True,
    )
    campaign = Campaign(indicators=[indicator], title="Title", description="Description", recommended_actions="Actions")

    workbook_bytes = export_campaign_to_excel_bytes(campaign)
    workbook = load_workbook(filename=BytesIO(workbook_bytes))
    worksheet = workbook.active

    assert worksheet[2][3].value == "BlockAndRemediate"
